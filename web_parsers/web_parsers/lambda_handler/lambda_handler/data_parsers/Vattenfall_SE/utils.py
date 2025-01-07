import tempfile
import time
from datetime import datetime
from datetime import datetime, timedelta

import openpyxl
import calendar
from playwright.sync_api import Page, sync_playwright

months = {
    "January": "januari",
    "February": "februari",
    "March": "mars",
    "April": "april",
    "May": "maj",
    "June": "juni",
    "July": "juli",
    "August": "augusti",
    "September": "september",
    "October": "oktober",
    "November": "november",
    "December": "december",
}


def login(page, username: str, password: str) -> None:
    """
    Logs into the Vattenfall Heat Business Portal with the provided credentials.

    Parameters:
    page (Page): The Playwright Page object used for browser interactions.
    username (str): The username for login.
    password (str): The password for login.

    Returns:
    None
    """
    login_url = "https://heatbusinessportal.vattenfall.se/Account/SignIn"
    page.goto(login_url)
    time.sleep(7)
    print(f"Logging in with username: {username}")
    page.fill("input#username", username)
    page.fill("input#password", password)
    page.click('button[type="submit"]')
    time.sleep(2)
    print("Login successful.")
    page.keyboard.press("Escape")
    time.sleep(5)


def get_point(page) -> str:
    """
    Retrieves the point value (Anläggnings-ID) from the given page.

    Parameters:
    page (Page): The Playwright Page object from which to retrieve the point value.

    Returns:
    str: The extracted point value.
    """
    try:
        accept_button = page.locator('button[class="ds-text ds-border ds-rounded-[20px] ds-px ds-py-xs"]')
        accept_button.click()
    except Exception as e:
        print("No button to accept cookies.")
        print(str(e))

    # Code Modification

    locator = page.locator('a.ds-hbp-menu-item[href="/report/consumption"]').nth(1)
    locator.scroll_into_view_if_needed()
    locator.click(force=True)

    # print("Anchor tag clicked successfully.")

    page.click('button[class = "ds-group ds-flex ds-justify-center ds-w-[62px] ds-min-h-[44px] ds-text-center ds-items-center md:ds-pl-xs md:ds-ml-xs md:ds-w-[40px]"]')
    # print("DropDown button click successfully")

    page.click('label[for = "HOUR"]')
    # print("Hour Click successfully")

def set_dates(page, date, key) -> None:
    """
    Sets the start and end dates for data extraction on the given page.

    Parameters:
    page (Page): The Playwright Page object on which to set the dates.
    start_date (datetime): The start date to be set.
    end_date (datetime): The end date to be set.

    Returns:
    None
    """

    position = 0 if key == 'start_date' else 1

    # Start Date
    if position == 0:
        page.click("span[data-testid='icon']")
        page.click("span.ds-icons-down")

        date_obj = datetime.strptime(date, "%Y-%m-%d")
        year = date_obj.year
        time.sleep(1)
        page.click(f"td:has-text('{year}')")

    else:
    # End Date
        page.locator("span[data-testid='icon']").nth(1).click()
        time.sleep(1)
        page.click("span.ds-icons-down")
        time.sleep(1)
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        year1 = date_obj.year

        if year1 <= 2024:
            year = date_obj.year
            month_number = 12
            user_date = 31
        else:
            current_date = datetime.now()
            year = current_date.year
            month_number = current_date.month
            user_date = current_date.day
            
        selector = f"td:has-text('{year}'):not(.ds-disabled)"  
        page.click(selector)
        time.sleep(1)
        page.locator("span[data-testid='icon']").nth(1).click()
        time.sleep(1)
        page.locator("span.ds-icons-down").nth(1).click()
        time.sleep(1)
        english_month = calendar.month_name[month_number]
        translated_month = months.get(english_month, None)
        if translated_month:
                # print(f"Checking for month: {english_month} ({translated_month})")
            page.locator('table.ds-calendar').wait_for(state='visible')
                # print("Months", month)
            month_button = page.locator(f"//td[contains(text(), '{translated_month}')]")
                # print("month_button",month_button)
            if month_button.count() > 0:
                month_button.first.click()
            else:
                print("Invalid month provided!")
        page.locator("span.ds-icons-calendar-01").nth(1).click()
        if year1 >= 2025:
            page.locator("span.ds-icons-calendar-01").nth(1).click()
        date_to_select = user_date
        selected_date = page.locator(f'td.ds-selected:has-text("{date_to_select}")')
        if selected_date.count() > 0:
            print(f"Date {date_to_select} is already selected.")
        else:
            page.locator(f'td:not(.ds-disabled):has-text("{date_to_select}")').nth(0).click()
            print(f"Date {date_to_select} selected successfully.")
        time.sleep(3)





        




def export_data(page: Page) -> str:
    """
    Initiates the data export process on the given page and handles the file download.

    Parameters:
    page (Page): The Playwright Page object on which to perform the export actions.

    Returns:
    str: The file path to the downloaded Excel file.
    """
    #Entry Button
    button = page.locator('button[data-track-as="5"]').nth(1)
    button.scroll_into_view_if_needed()
    button.wait_for(state="visible", timeout=5000)  
    button.click()

    # Another Download Button
    button2 = page.locator('button[class = "ds-button ds-tiny ds-no-arrow !ds-min-w-[130px] ds-pr-xs ds-pl"]').nth(1)
    button2.scroll_into_view_if_needed()
    button2.wait_for(state="visible", timeout=5000) 
    button2.click()
    print("successfully click")
    time.sleep(1)


    with tempfile.TemporaryDirectory() as temp_dir:
        download_path = "/tmp/consumption.xlsx"

        def handle_download(download):
            try:
                download.save_as(download_path)
                print(f"File saved successfully to: {download_path}")
            except Exception as e:
                print(f"Error saving file: {str(e)}")

        page.on("download", handle_download)

        return download_path


def read_file_and_transform(path: str, point: str, start_date: str, end_date: str) -> list[dict]:
    """
    Reads an Excel file and transforms the data into a structured format.

    Parameters:
    path (str): The file path to the Excel file.
    point (str): The point value to be included in the transformed data.
    start_date (str): Start date for filtering data (format: YYYY-MM-DD).
    end_date (str): End date for filtering data (format: YYYY-MM-DD).

    Returns:
    list[dict]: A list of dictionaries containing the transformed data.
    """

    print("Starting transformation process...")
    
    # Validate the date range
    current_date = datetime.now().date()
    start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()

    if end_dt > current_date:
        return {'error': "Invalid year. Dates cannot be in the future."}

    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    row_data = []
    data = []
    series = []

    for cells in ws.iter_rows():
        row_data.append([str(cell.value) if isinstance(cell.value, datetime) else cell.value for cell in cells])

    # Filter rows with valid data
    filtered_row_data = [row for row in row_data[2:] if any(cell is not None for cell in row)]
    if not filtered_row_data:
        return {'error': "No data found in this date"}

    for val in filtered_row_data:
        if val[0] and val[1] is not None:
            # Parse the date and time in the correct format
            dt = datetime.strptime(str(val[5]), "%Y-%m-%d %H:%M:%S")

            # Filter based on the date range
            if start_dt <= dt.date() <= end_dt:
                value = float(val[6]) * 3600
                formatted_value = round(value, 2)
                series.append(
                    {
                        "duration": "1H",
                        "end": f"{dt.isoformat()}+01:00",
                        "value": formatted_value,
                        "name": "thermal energy",
                        "unit": "MJ",
                    }
                )

    if not series:
        return {'error': "No data extracted for this date range"}

    data.append(
        {
            "point": point,
            "timezone": "Europe/Stockholm",
            "series": series,
            "parser": "Vattenfall_Heat_SE",
            "group": "sum",
        }
    )
    print("Transformation process completed.")
    return data



def extract(username: str, password: str, start_date: datetime, end_date: datetime) -> list[dict]:
    """
    Extracts consumption data for the specified meters within the given date range.

    Parameters:
    username (str): The username for login.
    password (str): The password for login.
    start_date (datetime): The start date for data extraction.
    end_date (datetime): The end date for data extraction.
    meters (list[str]): A list of meter IDs for which data is to be extracted.

    Returns:
    list[dict]: A list of dictionaries containing the extracted data.
    """
    path = "/tmp/consumption.xlsx"
    print("Starting extraction process...")
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--disable-gpu", "--single-process", "--incognito"])
            context = browser.new_context()
            page = context.new_page()
            login(page, username, password)
            point = get_point(page)
            

            #Start date
            key = 'start_date'
            set_dates(page, start_date, key)
            #End Date
            key = 'end_date'
            set_dates(page, end_date, key)

            

            export_data(page)
            page.wait_for_timeout(5000)

            browser.close()

        data = read_file_and_transform(path, point, start_date, end_date)
        if "error" in data:
            return data
        print(f"Successfully extracted the values.")
        return data
    except Exception as e:
        print(f"An error occurred during the extraction process: {e}")
        return []